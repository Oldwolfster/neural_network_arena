import sqlite3
from abc import abstractmethod

from openpyxl.reader.excel import load_workbook

from src.neuroForge_original import mgr
from tabulate import tabulate
from src.reports._ReportUtils import clean_multiline_string
import pandas as pd
import os

class BaseReport:
    """Base class for SQL-driven reports with reusable query execution and tabulation."""
    
    def __init__(self, *args):
        """
        Initialize the report with a reference to the database.
        :param RamDB: Database connection object
        """
        self.db = args[0]
        mgr.menu_active = False # Close menu





    def run_report(self):
        """
        This method is invoked when user selects this report from Report Menu
        It will print purpose above the report and what to look for underneath
        Need to add ability to enter parameters.
        """
        #if 'Purpose' is availble print it.
        if hasattr(self, 'purpose') and callable(self.purpose):
            print(clean_multiline_string(self.purpose()))

        #run report
        self.report_logic() #delegate to child class for report logic.

        #if 'What to look for' is availble print it.
        if hasattr(self, 'what_to_look_for') and callable(self.what_to_look_for):
            print(clean_multiline_string(self.what_to_look_for()))


    def run_sql_report(self, sql: str, params: list = None, filename: str = "SQL_Report.xlsx"):
        """
        Execute a SQL query, retrieve results, and export them to an Excel file.

        :param sql: SQL query string with optional placeholders.
        :param params: List of parameters for the SQL query.
        :param filename: Name of the Excel file to save.
        """
        #try:
        result_set = self.db.query(sql, params or [], as_dict=True)
        if not result_set:
            print("No data returned.")
            return

        # Open an existing Excel file
        script_dir = os.path.dirname(os.path.abspath(__file__))         #Get loc of this script
        file_path = os.path.join(script_dir, "_XL_Template.xlsx")  #add excel file name
        print(f"full_file_path={file_path}")

        wb = load_workbook(file_path)

        #print("Worksheets in file:", wb.sheetnames)

        ws = wb["GeneralTemplate"]  # 🟢 Select the Pre-Formatted Worksheet
        df = pd.DataFrame(result_set)

    # 🟢 **Write Field Headers to Row 2**
        for col_idx, col_name in enumerate(df.columns, start=1):  # Start at Excel column A (1-based)
            ws.cell(row=2, column=col_idx, value=col_name)  # Place headers in Row 2

        # 🟢 **Write Data Starting at Row 3**
        for row_idx, row in enumerate(df.itertuples(index=False), start=3):  # Data starts in Row 3
            for col_idx, value in enumerate(row, start=1):  # Start at Excel column A (1-based)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 🟢 Auto-Size Columns Based on Content
        from openpyxl.utils import get_column_letter
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_length = max(len(str(col_name)), *[len(str(row[col_idx-1])) for row in df.itertuples(index=False)])
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2  # Add padding


        # 🟢 Save Back to the Same File
        wb.save(file_path)

        # 🟢 (Optional) Open in Excel
        os.startfile(file_path)  # Only works on Windows
        ws.sheet_view.selection[0].activeCell = "A2"


        #with pd.ExcelWriter(full_file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        #    df.to_excel(writer, sheet_name="NewSheet", index=False)




    @abstractmethod
    def report_logic(self, *args )  :  #-> List[Tuple[Any, ...]]:
        pass

    def run_sql_reportInCmdWindow(self, sql: str, params: list = None, limit: int = 10):
        """
        Execute a SQL query, retrieve results, and display them in a tabulated format.
        
        :param sql: SQL query string with optional placeholders.
        :param params: List of parameters for the SQL query.
        :param limit: Number of records to display per batch.
        """
        try:
            result_set = self.db.query(sql, params or [], as_dict=True)
            if not result_set:
                print("No data returned.")
                return
            
            headers = result_set[0].keys()  # Extract column names from first record

            # Loop through the records in batches of `limit`
            for i in range(0, len(result_set), limit):
                batch = result_set[i:i + limit]
                print(tabulate(batch, headers="keys", tablefmt="fancy_grid"))

        except Exception as e:
            print(f"Error running report: {e}")




    def run_sql_reportxl(self, sql: str, params: list = None, filename: str = "SQL_Report.xlsx"):
        """
        Execute a SQL query, retrieve results, and export them to an Excel file.

        :param sql: SQL query string with optional placeholders.
        :param params: List of parameters for the SQL query.
        :param filename: Name of the Excel file to save.
        """
        #try:
        result_set = self.db.query(sql, params or [], as_dict=True)
        if not result_set:
            print("No data returned.")
            return

        # Convert result set to DataFrame
        df = pd.DataFrame(result_set)

        # Write to Excel
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Query Results")

            # Auto-adjust column widths
            worksheet = writer.sheets["Query Results"]
            for col_idx, col in enumerate(df.columns, 1):
                max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(64 + col_idx)].width = max_length

        print(f"✅ Report saved as '{filename}'")

        # Open the file automatically (like OLE Automation)
        import os
        os.system(f'start excel "{filename}"')

        #except Exception as e:
        #    print(f"❌ Error running report: {e}")



    """ Notes on future nice reports
    
    Key Objectives for a Better Multi-Neuron Report
Retain the insights from single-neuron reports (seeing the stabilization of weights, bias, and MAE).
Reduce clutter while still providing meaningful analysis.
Allow selective focus on neurons without losing global trends.
🔥 Recommended Approaches
Here are a few strategies to enhance readability and interpretation without sacrificing key insights.

1️⃣ Focus Mode: Select a Single Neuron for Detailed Analysis
📌 What it is:

Allow the user to select one neuron (e.g., via dropdown or hover) and plot only that neuron’s weights, bias, and MAE.
This mirrors the single-neuron report, keeping the insights clear.
📌 Why it's useful:

Keeps things simple.
Maintains the predictable MAE curve vs. weight stabilization trend without clutter.
Can compare different neurons individually if needed.
✅ Implementation:

Create an interactive dropdown list of neurons.
User selects one, and the report updates dynamically to show its trajectory over epochs.
2️⃣ Aggregate Trends with Confidence Bands (Mean & Variability)
📌 What it is:

Instead of showing every weight trajectory, calculate the mean and standard deviation of all neurons' weights.
Plot:
Mean weight trend as a solid line.
Shaded region (confidence band) to show how much variation exists.
📌 Why it's useful:

Reveals overall stabilization trends across the network.
Shows if certain neurons behave differently (outliers will be outside the confidence band).
Helps answer: "Are all neurons stabilizing in a similar pattern?"
✅ Implementation:

Compute mean weight per epoch and standard deviation.
Use Matplotlib's fill_between() to shade the confidence band.
Allow toggling between individual vs. aggregated view.
3️⃣ Weight Change Rate Instead of Absolute Weights
📌 What it is:

Instead of plotting raw weight values (which clutter quickly), plot their rate of change per epoch.
This would show:
High updates early (large adjustments).
Diminishing updates as learning stabilizes.
📌 Why it's useful:

Helps answer: "When does learning slow down?"
Removes the issue of overlapping weights by focusing on relative change.
Works well for networks with many neurons.
✅ Implementation:

Compute first derivative (difference between consecutive epochs).
Plot rate of change over time instead of raw values.
4️⃣ Compare MAE vs. Total Weight Movement (Sum of Changes Per Epoch)
📌 What it is:

Instead of showing individual neurons' weights, track:
Total sum of weight changes per epoch (magnitude of updates across all neurons).
Compare this against MAE, which should show a diminishing return pattern.
📌 Why it's useful:

Reveals how global learning slows down over epochs.
Answers "When do weight updates stop making a meaningful impact?"
Simplifies the report without losing insight.
✅ Implementation:

Compute Σ |Δweight| (sum of absolute weight changes).
Plot this against MAE to show learning progression.
    
    Oscillation Detector (Gradient Direction Change Count)
Concept: Track how often a weight flips direction (i.e., from increasing to decreasing).
How? Count the number of times a weight crosses zero in its gradient:
If weight changes sign between epochs (e.g., +0.2 → -0.3 → +0.1), it’s oscillating.
A high oscillation count over epochs means the training is unstable.
✅ Graph Idea:

Bar chart: Number of times each neuron’s weights flipped sign.
Threshold line: If oscillation is too high, we flag it in red.
2️⃣ Weight Smoothing / Moving Average Comparison
Concept: Compute a smoothed weight trajectory (e.g., a rolling average over 10 epochs).
How?
If the smoothed version follows a clear trend but the raw version jumps around wildly → it's oscillating.
✅ Graph Idea:

Two lines for each neuron’s weight:
Raw weight updates (showing jagged oscillations).
Smoothed weight trajectory (highlighting the intended trend).
If they diverge heavily, we flag instability.
3️⃣ Weight Velocity Plot (Acceleration Indicator)
Concept: Instead of raw weights, plot the rate of change (Δ weight per epoch).
How?
If the weight velocity fluctuates wildly, the model is likely oscillating.
If velocity is consistently near zero, training has plateaued.
If velocity is too high, weights may be exploding.
✅ Graph Idea:

Histogram or line chart of weight velocity.
Zones:
Red for high oscillation (rapid up-down shifts).
Yellow for instability.
Green for smooth convergence.
4️⃣ MAE vs. Weight Stability Heatmap
Concept: Combine weight stability with MAE.
How?
Create a 2D heatmap with:
X-axis = Epochs
Y-axis = Weight Change Rate (stability metric)
If instability increases as MAE decreases → learning rate is too high.
If instability is low but MAE isn't improving → learning rate might be too low.
✅ Graph Idea:

A heatmap where darker areas indicate high instability.
Overlay MAE trend to correlate instability with loss improvem
    
    
    

import win32com.client

def close_excel(file_path):
    try:
        # Connect to Excel
        excel = win32com.client.Dispatch("Excel.Application")
        
        # Loop through all open workbooks
        for wb in excel.Workbooks:
            if wb.FullName.lower() == file_path.lower():
                wb.Close(SaveChanges=False)  # ✅ Force close without save prompt
                print(f"Closed: {file_path}")
                break  # Stop after closing the right file

        # Quit Excel if no more workbooks are open
        if excel.Workbooks.Count == 0:
            excel.Quit()
            print("Excel closed successfully.")

    except Exception as e:
        print(f"Error closing Excel: {e}")

# 🔹 Example Usage
file_path = r"C:\SynologyDrive\Development\PycharmProjects\neural_network_arena\src\reports\_XL_Template.xlsx"
close_excel(file_path)


"""